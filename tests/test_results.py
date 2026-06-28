import tempfile
import unittest
from pathlib import Path
from openpyxl import load_workbook

from toetsanalyse.database import SubjectDatabase
from toetsanalyse.results import (
    RESULT_STATUSES,
    export_scores_xlsx,
    ResultValidationError,
    parse_score,
    regrade_multiple_choice_question,
    save_score,
    save_status,
    score_limit_conflicts,
    stored_results,
    test_students,
)


class ResultsEntryTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.database = SubjectDatabase.create(
            Path(self.temporary_directory.name) / "natuurkunde.db", "Natuurkunde", "2025-2026"
        )
        self.year_id = self.database.scalar("SELECT id FROM school_years LIMIT 1")
        self.database.execute(
            "INSERT INTO classes(school_year_id, name) VALUES(?, ?)",
            (self.year_id, "Cluster A"),
        )
        self.database.execute(
            "INSERT INTO classes(school_year_id, name) VALUES(?, ?)",
            (self.year_id, "Cluster B"),
        )
        self.class_a = self.database.scalar("SELECT id FROM classes WHERE name='Cluster A'")
        self.class_b = self.database.scalar("SELECT id FROM classes WHERE name='Cluster B'")
        self.database.execute("INSERT INTO students(display_name) VALUES(?)", ("Mila",))
        self.database.execute("INSERT INTO students(display_name) VALUES(?)", ("Sam",))
        self.mila = self.database.scalar("SELECT id FROM students WHERE display_name='Mila'")
        self.sam = self.database.scalar("SELECT id FROM students WHERE display_name='Sam'")
        self.database.execute(
            "INSERT INTO enrollments(student_id, class_id, school_year_id) VALUES(?, ?, ?)",
            (self.mila, self.class_a, self.year_id),
        )
        self.database.execute(
            "INSERT INTO enrollments(student_id, class_id, school_year_id) VALUES(?, ?, ?)",
            (self.sam, self.class_b, self.year_id),
        )
        self.database.execute(
            "INSERT INTO tests(school_year_id, name, period, test_type) VALUES(?, ?, ?, ?)",
            (self.year_id, "SO Energie", "Toetsweek 1", "SO"),
        )
        self.test_id = self.database.scalar("SELECT id FROM tests WHERE name='SO Energie'")
        self.database.execute(
            "INSERT INTO test_classes(test_id, class_id) VALUES(?, ?)",
            (self.test_id, self.class_a),
        )
        self.database.execute(
            "INSERT INTO matrix_questions(test_id, question_number, maximum_score) VALUES(?, ?, ?)",
            (self.test_id, "1", 4),
        )
        self.database.execute(
            "INSERT INTO matrix_questions(test_id, question_number, maximum_score) VALUES(?, ?, ?)",
            (self.test_id, "2", 6),
        )
        self.question_ids = [
            row["id"]
            for row in self.database.rows("SELECT id FROM matrix_questions WHERE test_id=? ORDER BY id", (self.test_id,))
        ]

    def tearDown(self) -> None:
        self.database.close()
        self.temporary_directory.cleanup()

    def test_only_students_from_test_groups_are_available_for_entry(self) -> None:
        self.assertEqual(["Mila"], [student["display_name"] for student in test_students(self.database, self.test_id)])

    def test_extra_student_link_is_available_for_entry(self) -> None:
        self.database.execute(
            "INSERT INTO test_students(test_id, student_id) VALUES(?, ?)",
            (self.test_id, self.sam),
        )

        available = [student["display_name"] for student in test_students(self.database, self.test_id)]

        self.assertEqual(["Mila", "Sam"], available)

    def test_students_are_sorted_by_last_name(self) -> None:
        self.database.execute("UPDATE students SET first_name=?, last_name=? WHERE id=?", ("Mila", "de Vries", self.mila))
        self.database.execute("UPDATE students SET first_name=?, last_name=? WHERE id=?", ("Sam", "Aalbers", self.sam))
        self.database.execute("INSERT INTO students(display_name, first_name, last_name) VALUES(?, ?, ?)", ("Jordi", "Jordi", "van Roij"))
        jordi = self.database.scalar("SELECT id FROM students WHERE display_name='Jordi'")
        self.database.execute(
            "INSERT INTO enrollments(student_id, class_id, school_year_id) VALUES(?, ?, ?)",
            (jordi, self.class_b, self.year_id),
        )
        self.database.execute("INSERT INTO test_classes(test_id, class_id) VALUES(?, ?)", (self.test_id, self.class_b))

        ordered = [student["display_name"] for student in test_students(self.database, self.test_id)]

        self.assertEqual(["Sam", "Jordi", "Mila"], ordered)

    def test_score_validation_depends_on_half_point_setting(self) -> None:
        self.assertEqual(2.0, parse_score("2", 4))
        self.assertEqual(2.5, parse_score("2,5", 4, allow_half_points=True))
        with self.assertRaises(ResultValidationError):
            parse_score("2,5", 4)
        with self.assertRaises(ResultValidationError):
            parse_score("4,5", 4)
        with self.assertRaises(ResultValidationError):
            parse_score("-1", 4)

    def test_entered_scores_create_attempt_and_update_total(self) -> None:
        save_score(self.database, self.test_id, self.mila, self.question_ids[0], "2,5", allow_half_points=True)
        total = save_score(self.database, self.test_id, self.mila, self.question_ids[1], "5", allow_half_points=True)
        attempts, scores, _ = stored_results(self.database, self.test_id)

        self.assertEqual(7.5, total)
        self.assertEqual("gemaakt", attempts[self.mila]["status"])
        self.assertEqual(7.5, attempts[self.mila]["total_score"])
        self.assertEqual(2.5, scores[(self.mila, self.question_ids[0])])

    def test_score_limit_conflicts_find_scores_above_new_maximum(self) -> None:
        save_score(self.database, self.test_id, self.mila, self.question_ids[0], "4")
        save_score(self.database, self.test_id, self.sam, self.question_ids[0], "2")

        conflicts = score_limit_conflicts(self.database, self.test_id, self.question_ids[0], 3)

        self.assertEqual(1, conflicts["count"])
        self.assertEqual(4.0, conflicts["highest_score"])
        self.assertEqual([{"display_name": "Mila", "score": 4.0}], conflicts["examples"])

    def test_score_limit_conflicts_allow_equal_or_lower_scores(self) -> None:
        save_score(self.database, self.test_id, self.mila, self.question_ids[0], "3")

        conflicts = score_limit_conflicts(self.database, self.test_id, self.question_ids[0], 3)

        self.assertEqual({"count": 0, "highest_score": None, "examples": []}, conflicts)

    def test_absent_status_clears_existing_scores_and_blocks_entry(self) -> None:
        save_score(self.database, self.test_id, self.mila, self.question_ids[0], "2")
        save_status(self.database, self.test_id, self.mila, "absent")
        attempts, scores, _ = stored_results(self.database, self.test_id)

        self.assertEqual("absent", attempts[self.mila]["status"])
        self.assertIsNone(attempts[self.mila]["total_score"])
        self.assertNotIn((self.mila, self.question_ids[0]), scores)
        with self.assertRaises(ResultValidationError):
            save_score(self.database, self.test_id, self.mila, self.question_ids[1], "3")

    def test_not_analyze_status_keeps_existing_scores_but_blocks_new_entry(self) -> None:
        save_score(self.database, self.test_id, self.mila, self.question_ids[0], "2")
        save_status(self.database, self.test_id, self.mila, "niet analyseren")
        attempts, scores, _ = stored_results(self.database, self.test_id)

        self.assertEqual("niet analyseren", attempts[self.mila]["status"])
        self.assertEqual(2.0, attempts[self.mila]["total_score"])
        self.assertEqual(2.0, scores[(self.mila, self.question_ids[0])])
        with self.assertRaises(ResultValidationError):
            save_score(self.database, self.test_id, self.mila, self.question_ids[1], "3")

    def test_status_list_contains_onregelmatigheid(self) -> None:
        self.assertIn("onregelmatigheid", RESULT_STATUSES)

    def test_status_list_contains_not_analyze(self) -> None:
        self.assertIn("niet analyseren", RESULT_STATUSES)

    def test_multiple_choice_score_is_automatically_checked_with_one_correct_answer(self) -> None:
        self.database.execute(
            "INSERT INTO matrix_questions(test_id, question_number, maximum_score, is_multiple_choice, multiple_choice_answer) "
            "VALUES(?, ?, ?, 1, ?)",
            (self.test_id, "3", 2, "E"),
        )
        question_id = self.database.scalar("SELECT id FROM matrix_questions WHERE question_number='3'")

        total_after_wrong = save_score(self.database, self.test_id, self.mila, question_id, "A")
        total_after_right = save_score(self.database, self.test_id, self.mila, question_id, "E")
        attempts, scores, responses = stored_results(self.database, self.test_id)

        self.assertEqual(0.0, total_after_wrong)
        self.assertEqual(2.0, total_after_right)
        self.assertEqual(2.0, scores[(self.mila, question_id)])
        self.assertEqual("E", responses[(self.mila, question_id)])
        self.assertEqual(2.0, attempts[self.mila]["total_score"])

    def test_multiple_choice_not_made_response_counts_as_zero_total(self) -> None:
        self.database.execute(
            "INSERT INTO matrix_questions(test_id, question_number, maximum_score, is_multiple_choice, multiple_choice_answer) "
            "VALUES(?, ?, ?, 1, ?)",
            (self.test_id, "3", 2, "E"),
        )
        question_id = self.database.scalar("SELECT id FROM matrix_questions WHERE question_number='3'")

        total = save_score(self.database, self.test_id, self.mila, question_id, "N")
        attempts, scores, responses = stored_results(self.database, self.test_id)

        self.assertEqual(0.0, total)
        self.assertEqual(0.0, scores[(self.mila, question_id)])
        self.assertEqual("N", responses[(self.mila, question_id)])
        self.assertEqual(0.0, attempts[self.mila]["total_score"])

    def test_multiple_choice_regrade_runs_after_answer_key_change(self) -> None:
        self.database.execute(
            "INSERT INTO matrix_questions(test_id, question_number, maximum_score, is_multiple_choice, multiple_choice_answer) "
            "VALUES(?, ?, ?, 1, ?)",
            (self.test_id, "4", 3, "B"),
        )
        question_id = self.database.scalar("SELECT id FROM matrix_questions WHERE question_number='4'")
        save_score(self.database, self.test_id, self.mila, question_id, "B")
        save_score(self.database, self.test_id, self.sam, question_id, "D")
        self.database.execute(
            "UPDATE matrix_questions SET multiple_choice_answer=? WHERE id=?",
            ("D", question_id),
        )

        changed = regrade_multiple_choice_question(self.database, self.test_id, question_id)
        attempts, scores, _ = stored_results(self.database, self.test_id)

        self.assertEqual(2, changed)
        self.assertEqual(0.0, scores[(self.mila, question_id)])
        self.assertEqual(3.0, scores[(self.sam, question_id)])
        self.assertEqual(0.0, attempts[self.mila]["total_score"])
        self.assertEqual(3.0, attempts[self.sam]["total_score"])

    def test_multiple_choice_correction_can_neutralize_question(self) -> None:
        self.database.execute(
            "INSERT INTO matrix_questions(test_id, question_number, maximum_score, is_multiple_choice, multiple_choice_answer, "
            "multiple_choice_correction_enabled, multiple_choice_correction_mode) "
            "VALUES(?, ?, ?, 1, ?, 1, 'neutralize')",
            (self.test_id, "5", 4, "C"),
        )
        question_id = self.database.scalar("SELECT id FROM matrix_questions WHERE question_number='5'")

        save_score(self.database, self.test_id, self.mila, question_id, "A")
        save_score(self.database, self.test_id, self.sam, question_id, "F")
        _attempts, scores, _responses = stored_results(self.database, self.test_id)

        self.assertEqual(4.0, scores[(self.mila, question_id)])
        self.assertEqual(4.0, scores[(self.sam, question_id)])

    def test_multiple_choice_correction_accepts_multiple_extra_options(self) -> None:
        self.database.execute(
            "INSERT INTO matrix_questions(test_id, question_number, maximum_score, is_multiple_choice, multiple_choice_answer, "
            "multiple_choice_correction_enabled, multiple_choice_correction_mode, multiple_choice_extra_answers) "
            "VALUES(?, ?, ?, 1, ?, 1, 'extra', ?)",
            (self.test_id, "6", 2, "A", "B,D,F"),
        )
        question_id = self.database.scalar("SELECT id FROM matrix_questions WHERE question_number='6'")

        save_score(self.database, self.test_id, self.mila, question_id, "B")
        save_score(self.database, self.test_id, self.sam, question_id, "D")
        self.database.execute("INSERT INTO students(display_name) VALUES(?)", ("Noor",))
        noor = self.database.scalar("SELECT id FROM students WHERE display_name='Noor'")
        self.database.execute(
            "INSERT INTO enrollments(student_id, class_id, school_year_id) VALUES(?, ?, ?)",
            (noor, self.class_a, self.year_id),
        )
        save_score(self.database, self.test_id, noor, question_id, "F")
        self.database.execute("INSERT INTO students(display_name) VALUES(?)", ("Iris",))
        iris = self.database.scalar("SELECT id FROM students WHERE display_name='Iris'")
        self.database.execute(
            "INSERT INTO enrollments(student_id, class_id, school_year_id) VALUES(?, ?, ?)",
            (iris, self.class_a, self.year_id),
        )
        save_score(self.database, self.test_id, iris, question_id, "C")
        _attempts, scores, _responses = stored_results(self.database, self.test_id)

        self.assertEqual(2.0, scores[(self.mila, question_id)])
        self.assertEqual(2.0, scores[(self.sam, question_id)])
        self.assertEqual(2.0, scores[(noor, question_id)])
        self.assertEqual(0.0, scores[(iris, question_id)])

    def test_score_export_contains_scores_per_student_and_question_in_question_order(self) -> None:
        self.database.execute(
            "INSERT INTO matrix_questions(test_id, question_number, subquestion, maximum_score) VALUES(?, ?, ?, ?)",
            (self.test_id, "1", "a", 2),
        )
        question_1a = self.database.scalar(
            "SELECT id FROM matrix_questions WHERE test_id=? AND question_number='1' AND subquestion='a'",
            (self.test_id,),
        )
        save_score(self.database, self.test_id, self.mila, self.question_ids[0], "3")
        save_score(self.database, self.test_id, self.mila, question_1a, "1")
        save_score(self.database, self.test_id, self.mila, self.question_ids[1], "5")

        output = Path(self.temporary_directory.name) / "scores.xlsx"
        export_scores_xlsx(self.database, self.test_id, output)

        workbook = load_workbook(output, data_only=True)
        try:
            sheet = workbook["Scores"]
            headers = [cell.value for cell in sheet[4]]
            self.assertEqual(["Leerlingnummer", "Leerling", "Groep", "Status"], headers[:4])
            self.assertEqual(["1 / 4", "1a / 2", "2 / 6"], headers[4:7])
            self.assertEqual("Mila", sheet["B5"].value)
            self.assertEqual(3, sheet["E5"].value)
            self.assertEqual(1, sheet["F5"].value)
            self.assertEqual(5, sheet["G5"].value)
            self.assertEqual(9, sheet["H5"].value)
        finally:
            workbook.close()

    def test_score_export_keeps_non_made_students_without_score_values(self) -> None:
        self.database.execute("INSERT INTO students(display_name) VALUES(?)", ("Noor",))
        noor = self.database.scalar("SELECT id FROM students WHERE display_name='Noor'")
        self.database.execute(
            "INSERT INTO enrollments(student_id, class_id, school_year_id) VALUES(?, ?, ?)",
            (noor, self.class_a, self.year_id),
        )
        save_score(self.database, self.test_id, noor, self.question_ids[0], "3")
        save_status(self.database, self.test_id, noor, "absent")

        output = Path(self.temporary_directory.name) / "scores_absent.xlsx"
        export_scores_xlsx(self.database, self.test_id, output)

        workbook = load_workbook(output, data_only=True)
        try:
            sheet = workbook["Scores"]
            rows = list(sheet.iter_rows(min_row=5, values_only=True))
            noor_row = next(row for row in rows if row[1] == "Noor")
            self.assertEqual("absent", noor_row[3])
            self.assertIsNone(noor_row[4])
            self.assertIsNone(noor_row[5])
            self.assertIsNone(noor_row[6])
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
