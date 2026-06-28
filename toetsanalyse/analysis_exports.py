from __future__ import annotations

import html
from pathlib import Path
from typing import Any


DEFAULT_ANALYSIS_EXPORT_OPTIONS = {
    "summary": True,
    "item_analysis": True,
    "multiple_choice": True,
    "group_analysis": True,
    "participants": False,
}


def _fmt(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 4)
    return value


def _status_label(value: object) -> str:
    if isinstance(value, dict):
        return str(value.get("label") or value.get("level") or "")
    return ""


def _esc(value: object) -> str:
    return html.escape("" if value is None else str(value))


def _num(value: object, decimals: int = 1, suffix: str = "") -> str:
    if value is None:
        return "-"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return _esc(value)
    if number.is_integer() and decimals == 0:
        text = str(int(number))
    else:
        text = f"{number:.{decimals}f}".replace(".", ",")
    return f"{text}{suffix}"


def _pct(value: object) -> str:
    return _num(value, 0, "%")


def _status_class(status: object) -> str:
    if not isinstance(status, dict):
        return "neutral"
    level = str(status.get("level") or "").casefold()
    if level in {"good", "excellent"}:
        return "good"
    if level in {"attention", "warning", "medium", "fair"}:
        return "attention"
    if level in {"bad", "problem", "weak"}:
        return "bad"
    return "neutral"


def _metric_value(metric: object, decimals: int = 2) -> str:
    if not isinstance(metric, dict):
        return "-"
    return _num(metric.get("value"), decimals)


def _section(title: str, body: str, note: str = "") -> str:
    note_html = f'<p class="section-note">{_esc(note)}</p>' if note else ""
    return f"""
    <section class="section">
      <div class="section-head">
        <h2>{_esc(title)}</h2>
        {note_html}
      </div>
      {body}
    </section>
    """


def _subtest_scope_section(data: dict[str, Any]) -> str:
    context = data.get("subtest_context")
    if not isinstance(context, dict):
        return ""
    cards = [
        ("Scope", f"Deeltoets: {context.get('name')}", "Opknippingsdeel"),
        ("Vragen", context.get("question_count"), "gekoppelde vragen"),
        ("Deel %", context.get("mean_percentage"), "gemiddelde score van dit deel"),
        ("Totaal %", context.get("total_mean_percentage"), "gemiddelde score van de hele toets"),
        ("Verschil", context.get("difference_percentage"), "deel minus totaal"),
        ("Samenhang", context.get("correlation_with_total"), "correlatie met totaalscore"),
    ]
    card_html = "".join(
        f"""
        <div class="kpi">
          <span>{_esc(label)}</span>
          <strong>{_esc(_num(value, 2 if label in {"Deel %", "Totaal %", "Verschil", "Samenhang"} else 0))}</strong>
          <small>{_esc(note)}</small>
        </div>
        """
        for label, value, note in cards
    )
    return _section(
        "Opknippingsdeel",
        f'<div class="kpi-grid">{card_html}</div>',
        "Deze export gebruikt alleen de vragen van de geselecteerde deeltoets. De vergelijking met de totaaltoets staat apart in het rapport.",
    )


def _quality_card(title: str, metric: object, description: str = "") -> str:
    status = metric.get("status") if isinstance(metric, dict) else {}
    status_text = _status_label(status)
    extra = ""
    if title == "SEM" and isinstance(metric, dict):
        extra = f'<div class="metric-extra">{_num(metric.get("score_value"), 2)} punten</div>'
    return f"""
      <div class="metric-card">
        <div class="metric-title">{_esc(title)}</div>
        <div class="metric-value">{_metric_value(metric)}{extra}</div>
        <span class="pill {_status_class(status)}">{_esc(status_text)}</span>
        <div class="metric-advice">{_esc(status.get("advice") if isinstance(status, dict) else description)}</div>
      </div>
    """


def _bar(label: object, percentage: object, note: str = "", color_class: str = "") -> str:
    try:
        width = max(0.0, min(100.0, float(percentage or 0)))
    except (TypeError, ValueError):
        width = 0.0
    return f"""
      <div class="bar-row">
        <div class="bar-label">{_esc(label)}</div>
        <div class="bar-track"><div class="bar-fill {color_class}" style="width:{width:.2f}%"></div></div>
        <div class="bar-value">{_pct(percentage)}{_esc(note)}</div>
      </div>
    """


def _summary_section(data: dict[str, Any]) -> str:
    test = data.get("test", {})
    summary = data.get("summary", {})
    quality = data.get("quality", {})
    cards = [
        ("Leerlingen", data.get("participant_count"), "complete resultaten"),
        ("Vragen", data.get("question_count"), "vragen in analyse"),
        ("Max. score", data.get("maximum_score"), "punten"),
        ("Gem. score", summary.get("mean_score"), f"van {_num(data.get('maximum_score'), 0)}"),
        ("Mediaan", summary.get("median_score"), "middelste score"),
        ("Modus", summary.get("mode_score"), f"{summary.get('mode_count') or 0}x"),
    ]
    card_html = "".join(
        f"""
        <div class="kpi">
          <span>{_esc(label)}</span>
          <strong>{_num(value, 1 if label.startswith('Gem') else 0)}</strong>
          <small>{_esc(note)}</small>
        </div>
        """
        for label, value, note in cards
    )
    quality_html = "".join(
        [
            _quality_card("Cronbach alpha", quality.get("alpha")),
            _quality_card("Gem. P-waarde", quality.get("p_value")),
            _quality_card("Gem. Rit", quality.get("rit")),
            _quality_card("Gem. Rir", quality.get("rir")),
            _quality_card("SEM", quality.get("sem")),
        ]
    )
    info_rows = [
        ("Toets", test.get("name")),
        ("Schooljaar", test.get("school_year")),
        ("Periode", test.get("period")),
        ("Niveau", test.get("level")),
        ("Jaarlaag", test.get("grade_year")),
    ]
    info_html = "".join(f"<span><b>{_esc(label)}</b>{_esc(value or '-')}</span>" for label, value in info_rows)
    return _section(
        "Samenvatting",
        f'<div class="info-strip">{info_html}</div><div class="kpi-grid">{card_html}</div>'
        f'<h3>Toetskwaliteit in één oogopslag</h3><div class="quality-grid">{quality_html}</div>',
        "Kleurstatussen zijn signalen voor bespreking; ze vervangen geen inhoudelijke beoordeling van de vraag.",
    )


def _item_rows(items: list[dict[str, Any]], child: bool = False) -> str:
    rows = []
    for item in items:
        status = item.get("status", {}) if isinstance(item, dict) else {}
        label = item.get("display_label") or item.get("label")
        rows.append(
            f"""
            <tr class="{'child-row' if child else ''}">
              <td>{_esc(label)}</td>
              <td>{_esc(item.get('description') or '-')}</td>
              <td class="num">{_num(item.get('maximum_score'), 1)}</td>
              <td class="num">{_num(item.get('p_value'), 2)}</td>
              <td class="num">{_num(item.get('rit'), 2)}</td>
              <td class="num">{_num(item.get('rir'), 2)}</td>
              <td><span class="pill {_status_class(status)}">{_esc(_status_label(status) or '-')}</span></td>
              <td>{_esc(status.get('reason') if isinstance(status, dict) else '')}</td>
            </tr>
            """
        )
        children = item.get("children", []) if isinstance(item, dict) else []
        if children:
            rows.append(_item_rows(children, True))
    return "".join(rows)


def _item_analysis_section(data: dict[str, Any]) -> str:
    items = data.get("question_group_analysis") or data.get("item_analysis") or []
    table = f"""
      <div class="table-card">
        <table>
          <thead><tr><th>Vraag</th><th>Omschrijving</th><th>Max.</th><th>P</th><th>Rit</th><th>Rir</th><th>Status</th><th>Onderbouwing</th></tr></thead>
          <tbody>{_item_rows(list(items))}</tbody>
        </table>
      </div>
    """
    return _section(
        "Itemanalyse",
        table,
        "P-waarde gaat over moeilijkheid. Rit en Rir tonen of een vraag onderscheid maakt tussen sterke en zwakkere leerlingen.",
    )


def _multiple_choice_section(data: dict[str, Any]) -> str:
    mc = data.get("multiple_choice", {})
    items = list(mc.get("items", [])) if isinstance(mc, dict) else []
    if not items:
        return _section("Meerkeuzeanalyse", '<div class="empty">Deze toets bevat geen meerkeuzevragen.</div>')
    cards = []
    for item in items:
        responses = item.get("responses", [])
        total = sum(int(response.get("count") or 0) for response in responses) or 1
        bars = "".join(
            _bar(
                response.get("option"),
                int(response.get("count") or 0) / total * 100,
                f" · {int(response.get('count') or 0)}x",
                "accepted" if response.get("accepted") else "",
            )
            for response in responses
        )
        conclusion = item.get("conclusion", {}) if isinstance(item.get("conclusion"), dict) else {}
        accepted = ", ".join(str(option) for option in item.get("accepted_answers", [])) or "-"
        cards.append(
            f"""
            <article class="mc-card">
              <div class="mc-head">
                <h3>Vraag {_esc(item.get('label'))}</h3>
                <span class="pill {_status_class(conclusion)}">{_esc(conclusion.get('label') or 'Analyse')}</span>
              </div>
              <div class="mc-meta">
                <span>Sleutel: <b>{_esc(item.get('answer_key') or '-')}</b></span>
                <span>Goedgekeurd: <b>{_esc(accepted)}</b></span>
                <span>N-score: <b>{_num(item.get('not_made_count'), 0)}</b></span>
              </div>
              <div class="mini-metrics">
                <span>P {_num(item.get('p_value'), 2)}</span>
                <span>Rit {_num(item.get('rit'), 2)}</span>
                <span>Rir {_num(item.get('rir'), 2)}</span>
              </div>
              <div class="bars">{bars}</div>
              <p>{_esc(conclusion.get('text') or '')}</p>
            </article>
            """
        )
    return _section(
        "Meerkeuzeanalyse",
        f'<div class="mc-grid">{"".join(cards)}</div>',
        "N-scores tellen als 0 voor de totaalscore, maar niet als gekozen alternatief in de verdeling van antwoorden.",
    )


def _group_analysis_section(data: dict[str, Any]) -> str:
    dimensions = data.get("group_dimensions") or []
    if not dimensions:
        dimensions = [
            {"title": value.get("title", key), "entries": value.get("entries", [])}
            for key, value in (data.get("groups") or {}).items()
            if isinstance(value, dict) and value.get("entries")
        ]
    cards = []
    for dimension in dimensions:
        entries = list(dimension.get("entries", []))[:8]
        if not entries:
            continue
        bars = "".join(_bar(entry.get("name"), entry.get("percentage"), "", "") for entry in entries)
        strongest = max(entries, key=lambda entry: float(entry.get("percentage") or 0))
        weakest = min(entries, key=lambda entry: float(entry.get("percentage") or 0))
        cards.append(
            f"""
            <article class="dimension-card">
              <h3>{_esc(dimension.get('title'))}</h3>
              <div class="bars">{bars}</div>
              <div class="insight">Sterkst: <b>{_esc(strongest.get('name'))}</b> ({_pct(strongest.get('percentage'))}).
              Aandachtspunt: <b>{_esc(weakest.get('name'))}</b> ({_pct(weakest.get('percentage'))}).</div>
            </article>
            """
        )
    if not cards:
        return _section("Analyse per onderdeel", '<div class="empty">Er zijn geen ingevulde taxonomieën of classificaties gevonden.</div>')
    return _section("Analyse per onderdeel", f'<div class="dimension-grid">{"".join(cards)}</div>')


def _participants_section(data: dict[str, Any]) -> str:
    participants = list(data.get("participants", []))
    if not participants:
        return ""
    rows = "".join(
        f"""
        <tr>
          <td>{_esc(participant.get('name'))}</td>
          <td class="num">{_num(participant.get('total_score'), 1)}</td>
          <td class="num">{_pct(participant.get('score_percentage'))}</td>
          <td class="num">{_num(participant.get('grade'), 1)}</td>
          <td>{_esc(participant.get('rank_label') or participant.get('rank') or '-')}</td>
        </tr>
        """
        for participant in participants
    )
    return _section(
        "Deelnemersoverzicht",
        f"""
        <div class="table-card">
          <table><thead><tr><th>Leerling</th><th>Score</th><th>% punten</th><th>Cijfer</th><th>Positie</th></tr></thead><tbody>{rows}</tbody></table>
        </div>
        """,
        "Neem dit onderdeel alleen mee als het rapport intern gedeeld wordt.",
    )


def build_section_analysis_report_html(data: dict[str, Any], options: dict[str, bool]) -> str:
    test = data.get("test", {})
    sections: list[str] = []
    if data.get("subtest_context"):
        sections.append(_subtest_scope_section(data))
    if options.get("summary", False):
        sections.append(_summary_section(data))
    if options.get("item_analysis", False):
        sections.append(_item_analysis_section(data))
    if options.get("multiple_choice", False):
        sections.append(_multiple_choice_section(data))
    if options.get("group_analysis", False):
        sections.append(_group_analysis_section(data))
    if options.get("participants", False):
        sections.append(_participants_section(data))
    content = "\n".join(section for section in sections if section)
    return f"""<!doctype html>
<html lang="nl">
<head>
<meta charset="utf-8">
<title>Sectierapport toetsanalyse - {_esc(test.get('name') or 'Toets')}</title>
<style>
@page {{ size:A4 landscape; margin:10mm; }}
* {{ box-sizing:border-box; }}
body {{ margin:0; background:#f6f8fb; color:#071f42; font-family:"Inter","Segoe UI",Arial,sans-serif; font-size:12px; line-height:1.45; }}
.report {{ max-width:1280px; margin:0 auto; }}
.hero {{ padding:24px 28px; margin-bottom:16px; border-radius:22px; background:linear-gradient(135deg,#0b2a4d,#1a3c63); color:white; box-shadow:0 16px 34px rgba(15,36,66,.14); }}
.hero h1 {{ margin:0 0 8px; font-size:30px; letter-spacing:-.03em; }}
.hero p {{ margin:0; color:#dce8f8; font-size:14px; }}
.section {{ background:white; border:1px solid #e5eaf1; border-radius:18px; padding:20px; margin:0 0 16px; break-inside:avoid-page; page-break-inside:avoid; box-shadow:0 10px 24px rgba(15,36,66,.05); }}
.section-head {{ display:flex; align-items:flex-end; justify-content:space-between; gap:16px; margin-bottom:14px; }}
h2 {{ margin:0; font-size:19px; letter-spacing:-.02em; }}
h3 {{ margin:10px 0 10px; font-size:14px; }}
.section-note {{ max-width:700px; margin:0; color:#62718d; }}
.info-strip {{ display:grid; grid-template-columns:repeat(5,1fr); gap:8px; margin-bottom:12px; }}
.info-strip span,.kpi,.metric-card,.mc-card,.dimension-card,.empty {{ border:1px solid #e5eaf1; border-radius:14px; background:#fbfcff; padding:12px; }}
.info-strip b {{ display:block; color:#6b7890; font-size:10px; text-transform:uppercase; letter-spacing:.05em; }}
.kpi-grid {{ display:grid; grid-template-columns:repeat(6,1fr); gap:10px; margin-bottom:14px; }}
.kpi span,.metric-title {{ color:#64728b; font-size:11px; font-weight:650; }}
.kpi strong {{ display:block; margin-top:6px; font-size:22px; }}
.kpi small {{ color:#71809a; }}
.quality-grid {{ display:grid; grid-template-columns:repeat(5,1fr); gap:10px; }}
.metric-value {{ font-size:20px; font-weight:750; margin:6px 0; }}
.metric-extra {{ display:inline-block; margin-left:6px; font-size:11px; color:#64728b; font-weight:650; }}
.metric-advice {{ margin-top:7px; color:#62718d; font-size:10px; }}
.pill {{ display:inline-flex; align-items:center; border-radius:999px; padding:4px 9px; font-size:10px; font-weight:700; background:#eef3fb; color:#42516b; }}
.pill.good {{ background:#e8f7ee; color:#168243; }} .pill.attention {{ background:#fff5dd; color:#9b6100; }} .pill.bad {{ background:#fdeaea; color:#c62828; }}
.table-card {{ overflow:hidden; border:1px solid #e5eaf1; border-radius:14px; }}
table {{ width:100%; border-collapse:collapse; }}
th {{ background:#edf3fc; text-align:left; font-size:11px; padding:9px 10px; color:#31415c; }}
td {{ padding:8px 10px; border-top:1px solid #edf1f6; vertical-align:top; }}
tbody tr:nth-child(even) td {{ background:#fbfcff; }}
.child-row td:first-child {{ padding-left:26px; color:#4b5b75; }}
.num {{ text-align:right; white-space:nowrap; }}
.mc-grid {{ display:grid; grid-template-columns:repeat(2,1fr); gap:12px; }}
.mc-head {{ display:flex; justify-content:space-between; align-items:center; gap:10px; }}
.mc-head h3 {{ margin:0; font-size:15px; }}
.mc-meta,.mini-metrics {{ display:flex; flex-wrap:wrap; gap:8px 14px; color:#56657f; margin:9px 0; }}
.mini-metrics span {{ background:#f2f6fb; border-radius:10px; padding:5px 8px; font-weight:700; color:#24334f; }}
.bar-row {{ display:grid; grid-template-columns:135px 1fr 72px; align-items:center; gap:8px; margin:7px 0; }}
.bar-label {{ color:#34445f; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }}
.bar-track {{ height:10px; border-radius:999px; background:#edf2f8; overflow:hidden; }}
.bar-fill {{ height:100%; border-radius:999px; background:#5b6df6; }}
.bar-fill.accepted {{ background:#36a96b; }}
.bar-value {{ text-align:right; font-weight:700; color:#253650; white-space:nowrap; }}
.dimension-grid {{ display:grid; grid-template-columns:repeat(2,1fr); gap:12px; }}
.dimension-card h3 {{ margin-top:0; }}
.insight {{ margin-top:10px; padding:10px; border-radius:12px; background:#f6f8fd; color:#455671; }}
.empty {{ color:#62718d; }}
@media print {{
  body {{ background:white; }}
  .section,.mc-card,.dimension-card,.metric-card,.kpi {{ box-shadow:none; break-inside:avoid-page; page-break-inside:avoid; }}
}}
</style>
</head>
<body>
<main class="report">
  <header class="hero">
    <h1>Sectierapport toetsanalyse</h1>
    <p>{_esc(test.get('name') or 'Toets')} · {_esc(test.get('school_year') or '')} · {_esc(test.get('level') or '')} {_esc(test.get('grade_year') or '')}</p>
  </header>
  {content}
</main>
<script>window.__chartsReady = true;</script>
</body>
</html>"""


def export_analysis_xlsx(data: dict[str, Any], output_path: str | Path, options: dict[str, bool]) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as error:
        raise RuntimeError("Excel-export vereist het pakket openpyxl.") from error

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Toetsanalyse"
    title_fill = PatternFill("solid", fgColor="1A3C63")
    header_fill = PatternFill("solid", fgColor="EAF1FB")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(color="071F42", bold=True)

    def style_header(row_number: int) -> None:
        for cell in sheet[row_number]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def append_section(title: str) -> None:
        if sheet.max_row > 1:
            sheet.append([])
        sheet.append([title])
        row = sheet.max_row
        sheet.cell(row, 1).font = title_font
        sheet.cell(row, 1).fill = title_fill

    test = data.get("test", {})
    sheet.append([f"Toetsanalyse - {test.get('name', '')}"])
    sheet.cell(1, 1).font = title_font
    sheet.cell(1, 1).fill = title_fill

    if options.get("summary", False):
        append_section("Samenvatting")
        summary = data.get("summary", {})
        quality = data.get("quality", {})
        rows = [
            ("Toets", test.get("name", "")),
            ("Schooljaar", test.get("school_year", "")),
            ("Periode", test.get("period", "")),
            ("Niveau", test.get("level", "")),
            ("Jaarlaag", test.get("grade_year", "")),
            ("Aantal leerlingen", data.get("participant_count", "")),
            ("Aantal vragen", data.get("question_count", "")),
            ("Maximale score", data.get("maximum_score", "")),
            ("Gemiddelde score", summary.get("mean_score")),
            ("Mediaan score", summary.get("median_score")),
            ("Modus score", summary.get("mode_score")),
            ("Standaarddeviatie score", summary.get("sd_score")),
            ("Cronbach alpha", quality.get("alpha", {}).get("value") if isinstance(quality.get("alpha"), dict) else ""),
            ("SEM (%)", quality.get("sem", {}).get("value") if isinstance(quality.get("sem"), dict) else ""),
            ("SEM (punten)", quality.get("sem", {}).get("score_value") if isinstance(quality.get("sem"), dict) else ""),
        ]
        for label, value in rows:
            sheet.append([label, _fmt(value)])

    if options.get("item_analysis", False):
        append_section("Itemanalyse")
        sheet.append(["Vraag", "Omschrijving", "Max.", "P-waarde", "Rit", "Rir", "P-status", "Rit-status", "Rir-status", "Beslissing", "Onderbouwing"])
        style_header(sheet.max_row)
        for item in data.get("question_group_analysis") or data.get("item_analysis", []):
            status = item.get("status", {}) if isinstance(item, dict) else {}
            sheet.append(
                [
                    item.get("display_label") or item.get("label"),
                    item.get("description", ""),
                    _fmt(item.get("maximum_score")),
                    _fmt(item.get("p_value")),
                    _fmt(item.get("rit")),
                    _fmt(item.get("rir")),
                    _status_label(item.get("p_status")),
                    _status_label(item.get("rit_status")),
                    _status_label(item.get("rir_status")),
                    _status_label(status),
                    str(status.get("reason") or "") if isinstance(status, dict) else "",
                ]
            )
            for child in item.get("children", []) if isinstance(item, dict) else []:
                child_status = child.get("status", {})
                sheet.append(
                    [
                        f"  {child.get('display_label') or child.get('label')}",
                        child.get("description", ""),
                        _fmt(child.get("maximum_score")),
                        _fmt(child.get("p_value")),
                        _fmt(child.get("rit")),
                        _fmt(child.get("rir")),
                        _status_label(child.get("p_status")),
                        _status_label(child.get("rit_status")),
                        _status_label(child.get("rir_status")),
                        _status_label(child_status),
                        str(child_status.get("reason") or "") if isinstance(child_status, dict) else "",
                    ]
                )

    if options.get("multiple_choice", False):
        append_section("Meerkeuzeanalyse")
        sheet.append(["Vraag", "Sleutel", "Goedgekeurde opties", "Niet gemaakt", "P-waarde", "Rit", "Rir", "Conclusie"])
        style_header(sheet.max_row)
        for item in data.get("multiple_choice", {}).get("items", []):
            sheet.append(
                [
                    item.get("label"),
                    item.get("answer_key", ""),
                    ", ".join(str(option) for option in item.get("accepted_answers", [])),
                    item.get("not_made_count", 0),
                    _fmt(item.get("p_value")),
                    _fmt(item.get("rit")),
                    _fmt(item.get("rir")),
                    item.get("conclusion", {}).get("text", "") if isinstance(item.get("conclusion"), dict) else "",
                ]
            )
            sheet.append(["", "Alternatief", "Aantal", "Percentage", "Goedgekeurd"])
            for response in item.get("responses", []):
                sheet.append(
                    [
                        "",
                        response.get("option"),
                        response.get("count"),
                        _fmt(response.get("percentage")),
                        "ja" if response.get("accepted") else "nee",
                    ]
                )

    if options.get("group_analysis", False):
        append_section("Analyse per onderdeel")
        sheet.append(["Onderdeel", "Waarde", "% van punten", "Aantal vragen", "Max. punten"])
        style_header(sheet.max_row)
        for key, group in data.get("groups", {}).items():
            if not isinstance(group, dict):
                continue
            for entry in group.get("entries", []):
                sheet.append(
                    [
                        group.get("title", key),
                        entry.get("name"),
                        _fmt(entry.get("percentage")),
                        entry.get("question_count", ""),
                        _fmt(entry.get("maximum_points")),
                    ]
                )

    if options.get("participants", False):
        append_section("Deelnemers")
        sheet.append(["Leerling", "Score", "% van punten", "Cijfer", "Rang", "Gedeelde positie"])
        style_header(sheet.max_row)
        for participant in data.get("participants", []):
            sheet.append(
                [
                    participant.get("name"),
                    _fmt(participant.get("total_score")),
                    _fmt(participant.get("score_percentage")),
                    _fmt(participant.get("grade")),
                    participant.get("rank"),
                    participant.get("tied_count"),
                ]
            )

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_index in range(1, sheet.max_column + 1):
        width = max(
            12,
            min(42, max((len(str(sheet.cell(row, column_index).value or "")) for row in range(1, sheet.max_row + 1)), default=8) + 2),
        )
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output
