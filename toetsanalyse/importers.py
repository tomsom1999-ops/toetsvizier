from __future__ import annotations

import csv
import difflib
import logging
import sqlite3
from dataclasses import dataclass
from pathlib import Path

from .database import SubjectDatabase
from .results import QUESTION_ORDER_SQL, RESULT_STATUSES, save_score, save_status


MAGISTER_HEADERS = {
    "stamnummer": "student_number",
    "roepnaam": "first_name",
    "tussenvoegsel": "prefix",
    "achternaam": "last_name",
}

OPTIONAL_HEADERS = {
    "email": "email",
    "emailadres": "email",
    "emailadresleerling": "email",
}

LOGGER = logging.getLogger("toetsvizier")
LOGGER.propagate = False


def _log_exception(message: str) -> None:
    if LOGGER.handlers:
        LOGGER.exception(message)


class StudentImportError(ValueError):
    pass


class ResultImportError(ValueError):
    pass


@dataclass(frozen=True)
class ImportedStudent:
    student_number: str
    first_name: str
    last_name: str
    display_name: str
    email: str = ""


@dataclass(frozen=True)
class ImportResult:
    added: int
    updated: int
    skipped: int


@dataclass(frozen=True)
class ResultImportSummary:
    updated_scores: int
    cleared_scores: int
    status_updates: int
    skipped_rows: int


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header(value: object) -> str:
    return "".join(character for character in _text(value).lower() if character.isalnum())


def _is_summary_row(values: dict[str, str]) -> bool:
    first_value = values.get("student_number", "").lower()
    return first_value in {"totaal", "total"} or first_value.startswith("totaal ")


def read_magister_students(path: Path | str) -> tuple[list[ImportedStudent], int]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as error:
        raise StudentImportError(
            "Excel-import vereist het pakket openpyxl. Installeer de onderdelen uit requirements.txt."
        ) from error
    try:
        workbook = load_workbook(filename=Path(path), read_only=True, data_only=True)
    except Exception as error:
        raise StudentImportError(f"Het Excelbestand kan niet worden gelezen: {error}") from error
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        columns = None
        data_rows = None
        for row_number, row in enumerate(rows, start=1):
            possible_columns = {_header(value): index for index, value in enumerate(row)}
            if all(name in possible_columns for name in MAGISTER_HEADERS):
                columns = possible_columns
                data_rows = rows
                break
            if row_number >= 50:
                break
        if columns is None or data_rows is None:
            expected = ", ".join(MAGISTER_HEADERS)
            raise StudentImportError(
                f"Geen Magister-kopregel gevonden in de eerste 50 rijen. Verwacht: {expected}."
            )
        students: list[ImportedStudent] = []
        skipped = 0
        for row in data_rows:
            values = {
                field: _text(row[columns[heading]]) if columns[heading] < len(row) else ""
                for heading, field in MAGISTER_HEADERS.items()
            }
            for heading, field in OPTIONAL_HEADERS.items():
                if heading in columns:
                    values[field] = _text(row[columns[heading]]) if columns[heading] < len(row) else ""
            if not any(values.values()):
                continue
            if _is_summary_row(values):
                continue
            surname = " ".join(part for part in (values["prefix"], values["last_name"]) if part)
            display_name = " ".join(part for part in (values["first_name"], surname) if part)
            if not values["student_number"] or not display_name:
                skipped += 1
                continue
            students.append(
                ImportedStudent(
                    student_number=values["student_number"],
                    first_name=values["first_name"],
                    last_name=surname,
                    display_name=display_name,
                    email=values.get("email", ""),
                )
            )
        return students, skipped
    finally:
        workbook.close()


def import_students(
    database: SubjectDatabase,
    school_year_id: int,
    class_id: int,
    students: list[ImportedStudent],
    skipped: int = 0,
) -> ImportResult:
    added = 0
    updated = 0
    try:
        for student in students:
            existing = database.connection.execute(
                "SELECT id FROM students WHERE student_number_key=LOWER(TRIM(?)) ORDER BY id LIMIT 1",
                (student.student_number,),
            ).fetchone()
            if existing:
                student_id = existing["id"]
                database.connection.execute(
                    "UPDATE students SET display_name=?, student_number=?, first_name=?, last_name=?, email=?, "
                    "updated_at=CURRENT_TIMESTAMP WHERE id=?",
                    (
                        student.display_name,
                        student.student_number,
                        student.first_name,
                        student.last_name,
                        student.email,
                        student_id,
                    ),
                )
                enrollment = database.connection.execute(
                    "SELECT id FROM enrollments WHERE student_id=? AND school_year_id=? LIMIT 1",
                    (student_id, school_year_id),
                ).fetchone()
                if enrollment:
                    database.connection.execute(
                        "UPDATE enrollments SET class_id=? WHERE id=?",
                        (class_id, enrollment["id"]),
                    )
                else:
                    database.connection.execute(
                        "INSERT INTO enrollments(student_id, class_id, school_year_id) VALUES(?, ?, ?)",
                        (student_id, class_id, school_year_id),
                    )
                updated += 1
            else:
                cursor = database.connection.execute(
                    "INSERT INTO students(display_name, student_number, first_name, last_name, email) "
                    "VALUES(?, ?, ?, ?, ?)",
                    (
                        student.display_name,
                        student.student_number,
                        student.first_name,
                        student.last_name,
                        student.email,
                    ),
                )
                database.connection.execute(
                    "INSERT INTO enrollments(student_id, class_id, school_year_id) VALUES(?, ?, ?)",
                    (cursor.lastrowid, class_id, school_year_id),
                )
                added += 1
        database.connection.commit()
    except sqlite3.IntegrityError as error:
        database.connection.rollback()
        raise StudentImportError(
            "Een leerlingnummer komt al bij een andere leerling voor. "
            "Controleer dubbele leerlingnummers voordat u importeert."
        ) from error
    except Exception:
        database.connection.rollback()
        raise
    return ImportResult(added=added, updated=updated, skipped=skipped)


def _normalized_header(text: str) -> str:
    return "".join(character for character in text.lower() if character.isalnum())


def _validated_headers(raw_headers: list[str]) -> list[str]:
    headers = [header or f"Kolom {index + 1}" for index, header in enumerate(raw_headers)]
    seen: dict[str, str] = {}
    duplicates: list[str] = []
    for header in headers:
        normalized = _normalized_header(header)
        if not normalized:
            continue
        if normalized in seen:
            duplicates.append(f"{seen[normalized]} / {header}")
        else:
            seen[normalized] = header
    if duplicates:
        examples = ", ".join(duplicates[:5])
        extra = f" en {len(duplicates) - 5} meer" if len(duplicates) > 5 else ""
        raise ResultImportError(
            "Dubbele kolomnamen gevonden. Geef elke kolom een unieke naam voordat je importeert: "
            f"{examples}{extra}."
        )
    return headers


def _match_header(headers: list[str], aliases: list[str]) -> str | None:
    normalized_headers = {_normalized_header(header): header for header in headers}
    for alias in aliases:
        found = normalized_headers.get(_normalized_header(alias))
        if found:
            return found
    return None


def _closest_student_id(name_value: str, by_name: dict[str, int], cutoff: float = 0.84) -> int | None:
    if not name_value:
        return None
    match = difflib.get_close_matches(name_value.lower(), by_name.keys(), n=1, cutoff=cutoff)
    if not match:
        return None
    return by_name.get(match[0])


def _closest_student_name(name_value: str, by_name: dict[str, int], cutoff: float = 0.84) -> str | None:
    if not name_value:
        return None
    match = difflib.get_close_matches(name_value.lower(), by_name.keys(), n=1, cutoff=cutoff)
    if not match:
        return None
    return match[0]


def read_results_rows(path: Path | str) -> tuple[list[str], list[dict[str, str]]]:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".csv":
        try:
            with source.open("r", encoding="utf-8-sig", newline="") as handle:
                sample = handle.read(4096)
                handle.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(handle, dialect)
                raw_rows = [row for row in reader if any(_text(value) for value in row)]
        except Exception as error:
            _log_exception("CSV-resultatenbestand kon niet worden gelezen.")
            raise ResultImportError(f"Het CSV-bestand kan niet worden gelezen: {error}") from error
        if not raw_rows:
            raise ResultImportError("Het bestand bevat geen gegevens.")
        headers = _validated_headers([_text(value) for value in raw_rows[0]])
        rows = [{headers[index]: _text(row[index]) if index < len(row) else "" for index in range(len(headers))} for row in raw_rows[1:]]
        return headers, rows
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as error:
        raise ResultImportError(
            "Excel-import vereist het pakket openpyxl. Installeer de onderdelen uit requirements.txt."
        ) from error
    try:
        workbook = load_workbook(filename=source, read_only=True, data_only=True)
    except Exception as error:
        _log_exception("Excel-resultatenbestand kon niet worden gelezen.")
        raise ResultImportError(f"Het Excelbestand kan niet worden gelezen: {error}") from error
    try:
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        header_row = None
        for _ in range(50):
            candidate = next(iterator, None)
            if candidate is None:
                break
            values = [_text(cell) for cell in candidate]
            if sum(1 for value in values if value) >= 2:
                header_row = values
                break
        if not header_row:
            raise ResultImportError("Geen kolomkop gevonden in de eerste 50 rijen.")
        headers = _validated_headers(header_row)
        rows: list[dict[str, str]] = []
        for row in iterator:
            values = [_text(cell) for cell in row]
            if not any(values):
                continue
            rows.append({headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))})
        return headers, rows
    finally:
        workbook.close()


def suggest_results_mapping(headers: list[str], questions: list[dict]) -> dict[str, object]:
    mapping: dict[str, object] = {
        "student_number": _match_header(headers, ["leerlingnummer", "studentnummer", "stamnummer", "nummer"]),
        "student_name": _match_header(headers, ["naam", "leerling", "weergavenaam"]),
        "status": _match_header(headers, ["status", "leerlingstatus"]),
        "questions": {},
    }
    normalized_headers = {_normalized_header(header): header for header in headers}
    for question in questions:
        label = str(question["label"])
        candidates = [
            label,
            f"v{label}",
            f"vraag{label}",
            f"score{label}",
            f"q{label}",
        ]
        selected = None
        for candidate in candidates:
            selected = normalized_headers.get(_normalized_header(candidate))
            if selected:
                break
        mapping["questions"][question["id"]] = selected
    return mapping


def preview_results_import(
    database: SubjectDatabase,
    test_id: int,
    rows: list[dict[str, str]],
    mapping: dict[str, object],
) -> dict[str, object]:
    questions = {
        row["id"]: dict(row)
        for row in database.rows(
            "SELECT id, question_number || COALESCE(subquestion, '') AS label "
            f"FROM matrix_questions WHERE test_id=? ORDER BY {QUESTION_ORDER_SQL}",
            (test_id,),
        )
    }
    by_number = {}
    for row in database.rows("SELECT id, student_number_key, display_name FROM students"):
        if row["student_number_key"]:
            by_number[str(row["student_number_key"]).strip().lower()] = row["id"]
    by_name = {
        str(row["display_name"]).strip().lower(): row["id"]
        for row in database.rows("SELECT id, display_name FROM students")
    }
    unknown_students = 0
    matched_rows = 0
    score_cells = 0
    unmatched_students: list[str] = []
    fuzzy_matches: list[dict[str, str]] = []
    invalid_statuses: list[dict[str, str]] = []
    preview: list[dict[str, str]] = []
    for row_index, row in enumerate(rows, start=2):
        student_id = None
        number_column = mapping.get("student_number")
        name_column = mapping.get("student_name")
        status_column = mapping.get("status")
        number_value = _text(row.get(str(number_column))) if number_column else ""
        name_value = _text(row.get(str(name_column))) if name_column else ""
        raw_status = _text(row.get(str(status_column))) if status_column else ""
        if raw_status and raw_status.lower() not in RESULT_STATUSES:
            invalid_statuses.append(
                {
                    "row": str(row_index),
                    "student": number_value or name_value or "(onbekende leerling)",
                    "status": raw_status,
                }
            )
        if number_column:
            student_id = by_number.get(number_value.lower())
        if student_id is None and name_column:
            student_id = by_name.get(name_value.lower())
        if student_id is None and name_column and name_value:
            closest_name = _closest_student_name(name_value, by_name)
            if closest_name:
                student_id = by_name.get(closest_name)
                fuzzy_matches.append({"input": name_value, "matched_to": closest_name})
        if student_id is None:
            unknown_students += 1
            unknown_label = number_value or name_value or "(leeg)"
            if unknown_label not in unmatched_students:
                unmatched_students.append(unknown_label)
            if number_column or name_column:
                continue
        matched_rows += 1
        score_values = {}
        for question_id, column_name in dict(mapping.get("questions", {})).items():
            if not column_name:
                continue
            raw_value = _text(row.get(str(column_name)))
            if raw_value:
                score_cells += 1
            score_values[questions.get(int(question_id), {}).get("label", str(question_id))] = raw_value
        if len(preview) < 25:
            preview_row = {"leerling": number_value or name_value}
            preview_row.update(score_values)
            preview.append(preview_row)
    return {
        "matched_rows": matched_rows,
        "unknown_students": unknown_students,
        "score_cells": score_cells,
        "preview_rows": preview,
        "unmatched_students": unmatched_students,
        "fuzzy_matches": fuzzy_matches,
        "invalid_statuses": invalid_statuses,
    }


def import_results(
    database: SubjectDatabase,
    test_id: int,
    rows: list[dict[str, str]],
    mapping: dict[str, object],
    allow_half_points: bool = False,
) -> ResultImportSummary:
    by_number = {}
    for row in database.rows("SELECT id, student_number_key, display_name FROM students"):
        if row["student_number_key"]:
            by_number[str(row["student_number_key"]).strip().lower()] = row["id"]
    by_name = {
        str(row["display_name"]).strip().lower(): row["id"]
        for row in database.rows("SELECT id, display_name FROM students")
    }
    updated_scores = 0
    cleared_scores = 0
    status_updates = 0
    skipped_rows = 0
    fuzzy_overrides = {
        str(key).strip().lower(): str(value).strip().lower()
        for key, value in dict(mapping.get("fuzzy_overrides", {})).items()
        if str(key).strip() and str(value).strip()
    }
    try:
        for row in rows:
            student_id = None
            number_column = mapping.get("student_number")
            name_column = mapping.get("student_name")
            status_column = mapping.get("status")
            number_value = _text(row.get(str(number_column))) if number_column else ""
            name_value = _text(row.get(str(name_column))) if name_column else ""
            if number_column:
                student_id = by_number.get(number_value.lower())
            if student_id is None and name_column:
                student_id = by_name.get(name_value.lower())
            if student_id is None and name_column and name_value:
                override_target = fuzzy_overrides.get(name_value.lower())
                if override_target:
                    student_id = by_name.get(override_target)
            if student_id is None:
                skipped_rows += 1
                continue
            raw_status = _text(row.get(str(status_column))) if status_column else ""
            if raw_status:
                status = raw_status.lower()
                if status not in RESULT_STATUSES:
                    raise ResultImportError(f"Onbekende leerlingstatus '{raw_status}'. Controleer de statuskolom.")
                save_status(database, test_id, student_id, status, commit=False)
                status_updates += 1
            for question_id, column_name in dict(mapping.get("questions", {})).items():
                if not column_name:
                    continue
                raw_value = _text(row.get(str(column_name)))
                total = save_score(
                    database,
                    test_id,
                    student_id,
                    int(question_id),
                    raw_value,
                    allow_half_points=allow_half_points,
                    commit=False,
                )
                if raw_value:
                    updated_scores += 1
                elif total is not None:
                    cleared_scores += 1
        database.connection.commit()
    except Exception:
        _log_exception("Resultatenimport is teruggedraaid na een fout.")
        database.connection.rollback()
        raise
    return ResultImportSummary(
        updated_scores=updated_scores,
        cleared_scores=cleared_scores,
        status_updates=status_updates,
        skipped_rows=skipped_rows,
    )
