from __future__ import annotations

import math
from pathlib import Path

from .database import SubjectDatabase
from .name_sorting import student_sort_key


QUESTION_ORDER_SQL = "CAST(question_number AS INTEGER), question_number, COALESCE(subquestion, ''), id"

RESULT_STATUSES = [
    "gemaakt",
    "niet analyseren",
    "onregelmatigheid",
    "absent",
    "ziek",
    "geoorloofd afwezig",
    "ongeoorloofd afwezig",
    "vrijstelling",
    "niet gemaakt",
    "ongeldig",
]


class ResultValidationError(ValueError):
    pass


class ResultExportError(ValueError):
    pass


def normalize_multiple_choice_response(raw_value: str) -> str | None:
    value = raw_value.strip().upper()
    if not value:
        return None
    if len(value) != 1 or not value.isalpha():
        raise ResultValidationError("Voer bij meerkeuze precies een letter in (bijvoorbeeld A of E).")
    return value


def is_not_made_score(raw_value: str) -> bool:
    return raw_value.strip().casefold() == "n"


def normalize_multiple_choice_option_list(raw_value: str) -> list[str]:
    text = raw_value.strip()
    if not text:
        return []
    parts = [part.strip() for part in text.replace(";", ",").split(",")]
    options: list[str] = []
    seen: set[str] = set()
    for part in parts:
        if not part:
            continue
        option = normalize_multiple_choice_response(part)
        if option and option not in seen:
            seen.add(option)
            options.append(option)
    return options


def _multiple_choice_score(question: dict, response_text: str | None) -> float | None:
    if response_text is None:
        return None
    maximum = float(question["maximum_score"])
    answer_key = normalize_multiple_choice_response(str(question["multiple_choice_answer"] or ""))
    if answer_key is None:
        raise ResultValidationError("Voor deze meerkeuzevraag ontbreekt de antwoordsleutel.")
    correction_enabled = bool(question["multiple_choice_correction_enabled"]) if "multiple_choice_correction_enabled" in question.keys() else False
    correction_mode = str(question["multiple_choice_correction_mode"] or "none") if "multiple_choice_correction_mode" in question.keys() else "none"
    if correction_enabled and correction_mode == "neutralize":
        return maximum
    accepted = {answer_key}
    if correction_enabled and correction_mode == "extra":
        extra_answers = str(question["multiple_choice_extra_answers"] or "") if "multiple_choice_extra_answers" in question.keys() else ""
        accepted.update(normalize_multiple_choice_option_list(extra_answers))
    return maximum if response_text in accepted else 0.0


def parse_score(raw_value: str, maximum_score: float, allow_half_points: bool = False) -> float | None:
    value = raw_value.strip().replace(",", ".")
    if not value:
        return None
    try:
        score = float(value)
    except ValueError as error:
        raise ResultValidationError("Voer een numerieke score in.") from error
    if not math.isfinite(score):
        raise ResultValidationError("Voer een geldige numerieke score in.")
    if score < 0:
        raise ResultValidationError("Een score kan niet negatief zijn.")
    if score > maximum_score:
        raise ResultValidationError(f"De score mag niet hoger zijn dan {maximum_score:g}.")
    if allow_half_points:
        if not math.isclose(score * 2, round(score * 2), rel_tol=0.0, abs_tol=1e-9):
            raise ResultValidationError("Alleen hele of halve punten zijn toegestaan.")
    elif not math.isclose(score, round(score), rel_tol=0.0, abs_tol=1e-9):
        raise ResultValidationError("Alleen hele punten zijn toegestaan.")
    return score


def test_students(database: SubjectDatabase, test_id: int) -> list[dict]:
    students = [
        dict(row)
        for row in database.rows(
            "SELECT s.id, s.display_name, COALESCE(s.student_number, '') AS student_number, "
            "COALESCE(s.first_name, '') AS first_name, COALESCE(s.last_name, '') AS last_name, "
            "COALESCE(GROUP_CONCAT(DISTINCT c.name), '') AS groups, "
            "MAX(CASE WHEN ts.student_id IS NOT NULL THEN 1 ELSE 0 END) AS is_extra_student "
            "FROM students s "
            "LEFT JOIN tests t ON t.id=? "
            "LEFT JOIN enrollments e ON e.student_id=s.id AND e.school_year_id=t.school_year_id "
            "LEFT JOIN classes c ON c.id=e.class_id "
            "LEFT JOIN test_classes tc ON tc.class_id=c.id AND tc.test_id=t.id "
            "LEFT JOIN test_students ts ON ts.student_id=s.id AND ts.test_id=t.id "
            "WHERE tc.test_id IS NOT NULL OR ts.test_id IS NOT NULL "
            "GROUP BY s.id, s.display_name, s.student_number, s.last_name, s.first_name "
            "ORDER BY s.display_name",
            (test_id,),
        )
    ]
    for student in students:
        if student.get("is_extra_student") and not str(student.get("groups") or "").strip():
            student["groups"] = "Los gekoppeld"
    students.sort(key=lambda student: student_sort_key(student["display_name"], student["first_name"], student["last_name"]))
    return students


def test_questions(database: SubjectDatabase, test_id: int) -> list[dict]:
    return [
        dict(row)
        for row in database.rows(
            "SELECT id, question_number || COALESCE(subquestion, '') AS label, maximum_score, "
            "is_multiple_choice, multiple_choice_answer, multiple_choice_correction_enabled, "
            "multiple_choice_correction_mode, multiple_choice_extra_answers "
            f"FROM matrix_questions WHERE test_id=? ORDER BY {QUESTION_ORDER_SQL}",
            (test_id,),
        )
    ]


def export_scores_xlsx(database: SubjectDatabase, test_id: int, output_path: str | Path) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as error:
        raise ResultExportError(
            "Excel-export vereist het pakket openpyxl. Installeer de onderdelen uit requirements.txt."
        ) from error

    test_row = database.connection.execute(
        "SELECT t.name, t.period, t.test_type, t.level, t.grade_year, sy.name AS school_year "
        "FROM tests t JOIN school_years sy ON sy.id=t.school_year_id WHERE t.id=?",
        (test_id,),
    ).fetchone()
    if not test_row:
        raise ResultExportError("De geselecteerde toets bestaat niet meer.")

    questions = test_questions(database, test_id)
    students = test_students(database, test_id)
    attempts, scores, responses = stored_results(database, test_id)
    maximum = sum(float(question["maximum_score"]) for question in questions)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Scores"
    title_fill = PatternFill("solid", fgColor="1A3C63")
    header_fill = PatternFill("solid", fgColor="EAF1FB")
    locked_fill = PatternFill("solid", fgColor="FDECEC")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(color="071F42", bold=True)

    headers = ["Leerlingnummer", "Leerling", "Groep", "Status"]
    headers.extend(f"{question['label']} / {float(question['maximum_score']):g}" for question in questions)
    headers.append(f"Totaal / {maximum:g}")

    sheet.append([f"Scores - {test_row['name']}"])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    title_cell = sheet.cell(1, 1)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.append(
        [
            f"Vak: {database.meta('subject_name', database.path.stem)}",
            f"Schooljaar: {test_row['school_year']}",
            f"Periode: {test_row['period']}",
            f"Niveau: {test_row['level'] or '-'}",
            f"Jaarlaag: {test_row['grade_year'] or '-'}",
        ]
    )
    sheet.append([])
    sheet.append(headers)
    header_row = 4
    for cell in sheet[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 34

    for student in students:
        student_id = int(student["id"])
        attempt = attempts.get(student_id, {})
        status = attempt.get("status", "gemaakt")
        row = [
            student.get("student_number", ""),
            student.get("display_name", ""),
            student.get("groups", ""),
            status,
        ]
        for question in questions:
            value = scores.get((student_id, int(question["id"])))
            row.append(None if value is None else float(value))
        total = attempt.get("total_score")
        row.append(None if total is None else float(total))
        sheet.append(row)
        if status != "gemaakt":
            for cell in sheet[sheet.max_row]:
                cell.fill = locked_fill

    sheet.freeze_panes = "E5"
    sheet.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(4, sheet.max_row)}"
    for row in sheet.iter_rows(min_row=5, min_col=5, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=5, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
    widths = {
        1: 16,
        2: max(18, min(42, max((len(str(student.get("display_name", ""))) for student in students), default=12) + 3)),
        3: 18,
        4: 18,
    }
    for column_index in range(1, len(headers) + 1):
        if column_index in widths:
            width = widths[column_index]
        elif column_index == len(headers):
            width = 14
        else:
            width = 11
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


def stored_results(
    database: SubjectDatabase, test_id: int
) -> tuple[dict[int, dict], dict[tuple[int, int], float], dict[tuple[int, int], str]]:
    attempts = {
        row["student_id"]: dict(row)
        for row in database.rows(
            "SELECT id, student_id, status, total_score FROM test_attempts WHERE test_id=?",
            (test_id,),
        )
    }
    scores = {
        (row["student_id"], row["question_id"]): float(row["score"])
        for row in database.rows(
            "SELECT a.student_id, s.question_id, s.score FROM scores s "
            "JOIN test_attempts a ON a.id=s.attempt_id WHERE a.test_id=? AND s.score IS NOT NULL",
            (test_id,),
        )
    }
    responses = {
        (row["student_id"], row["question_id"]): str(row["response_text"]).upper()
        for row in database.rows(
            "SELECT a.student_id, s.question_id, s.response_text FROM scores s "
            "JOIN test_attempts a ON a.id=s.attempt_id "
            "WHERE a.test_id=? AND s.response_text IS NOT NULL AND TRIM(s.response_text)<>''",
            (test_id,),
        )
    }
    return attempts, scores, responses


def score_limit_conflicts(
    database: SubjectDatabase,
    test_id: int,
    question_id: int,
    maximum_score: float,
) -> dict[str, object]:
    count = int(
        database.scalar(
            "SELECT COUNT(*) FROM scores sc "
            "JOIN test_attempts a ON a.id=sc.attempt_id "
            "WHERE a.test_id=? AND sc.question_id=? AND sc.score IS NOT NULL AND sc.score>?",
            (test_id, question_id, float(maximum_score)),
        )
        or 0
    )
    if not count:
        return {"count": 0, "highest_score": None, "examples": []}
    highest_score = database.scalar(
        "SELECT MAX(sc.score) FROM scores sc "
        "JOIN test_attempts a ON a.id=sc.attempt_id "
        "WHERE a.test_id=? AND sc.question_id=? AND sc.score IS NOT NULL AND sc.score>?",
        (test_id, question_id, float(maximum_score)),
    )
    examples = [
        dict(row)
        for row in database.rows(
            "SELECT s.display_name, sc.score FROM scores sc "
            "JOIN test_attempts a ON a.id=sc.attempt_id "
            "JOIN students s ON s.id=a.student_id "
            "WHERE a.test_id=? AND sc.question_id=? AND sc.score IS NOT NULL AND sc.score>? "
            "ORDER BY sc.score DESC, s.display_name LIMIT 5",
            (test_id, question_id, float(maximum_score)),
        )
    ]
    return {"count": count, "highest_score": highest_score, "examples": examples}


def save_score(
    database: SubjectDatabase,
    test_id: int,
    student_id: int,
    question_id: int,
    raw_value: str,
    allow_half_points: bool = False,
    *,
    commit: bool = True,
) -> float | None:
    current_attempt = database.connection.execute(
        "SELECT id, status FROM test_attempts WHERE test_id=? AND student_id=?",
        (test_id, student_id),
    ).fetchone()
    current_status = str(current_attempt["status"]) if current_attempt else ""
    if current_attempt and current_status != "gemaakt":
        raise ResultValidationError("Voor deze leerlingstatus kunnen geen scores worden ingevoerd.")
    row = database.rows(
        "SELECT maximum_score, is_multiple_choice, multiple_choice_answer, "
        "multiple_choice_correction_enabled, multiple_choice_correction_mode, multiple_choice_extra_answers "
        "FROM matrix_questions WHERE id=? AND test_id=?",
        (question_id, test_id),
    )
    if not row:
        raise ResultValidationError("Deze vraag hoort niet meer bij de geselecteerde toets.")
    question = row[0]
    maximum = float(question["maximum_score"])
    is_multiple_choice = bool(question["is_multiple_choice"])
    response_text = None
    if is_not_made_score(raw_value):
        response_text = "N"
        score = 0.0
    elif is_multiple_choice:
        response_text = normalize_multiple_choice_response(raw_value)
        score = _multiple_choice_score(question, response_text)
    else:
        score = parse_score(raw_value, maximum, allow_half_points=allow_half_points)
    connection = database.connection
    try:
        if score is not None:
            attempt = connection.execute(
                "SELECT id FROM test_attempts WHERE test_id=? AND student_id=?",
                (test_id, student_id),
            ).fetchone()
            if attempt is None:
                connection.execute(
                    "INSERT INTO test_attempts(test_id, student_id, status) VALUES(?, ?, 'gemaakt')",
                    (test_id, student_id),
                )
                attempt = connection.execute(
                    "SELECT id FROM test_attempts WHERE test_id=? AND student_id=?",
                    (test_id, student_id),
                ).fetchone()
            attempt_id = attempt["id"]
            if current_status != "niet analyseren":
                connection.execute(
                    "INSERT INTO test_attempts(test_id, student_id, status) VALUES(?, ?, 'gemaakt') "
                    "ON CONFLICT(test_id, student_id) DO UPDATE SET status='gemaakt', updated_at=CURRENT_TIMESTAMP",
                    (test_id, student_id),
                )
            connection.execute(
                "INSERT INTO scores(attempt_id, question_id, score, response_text) VALUES(?, ?, ?, ?) "
                "ON CONFLICT(attempt_id, question_id) DO UPDATE SET "
                "score=excluded.score, response_text=excluded.response_text",
                (attempt_id, question_id, score, response_text),
            )
        else:
            attempt = connection.execute(
                "SELECT id FROM test_attempts WHERE test_id=? AND student_id=?",
                (test_id, student_id),
            ).fetchone()
            if attempt:
                connection.execute(
                    "DELETE FROM scores WHERE attempt_id=? AND question_id=?",
                    (attempt["id"], question_id),
                )
        attempt = connection.execute(
            "SELECT id FROM test_attempts WHERE test_id=? AND student_id=?",
            (test_id, student_id),
        ).fetchone()
        if not attempt:
            if commit:
                connection.commit()
            return None
        total = connection.execute(
            "SELECT SUM(score) FROM scores WHERE attempt_id=?",
            (attempt["id"],),
        ).fetchone()[0]
        connection.execute(
            "UPDATE test_attempts SET total_score=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (total, attempt["id"]),
        )
        if commit:
            connection.commit()
        return float(total) if total is not None else None
    except Exception:
        if commit:
            connection.rollback()
        raise


def regrade_multiple_choice_question(
    database: SubjectDatabase, test_id: int, question_id: int
) -> int:
    question_rows = database.rows(
        "SELECT maximum_score, is_multiple_choice, multiple_choice_answer, "
        "multiple_choice_correction_enabled, multiple_choice_correction_mode, multiple_choice_extra_answers "
        "FROM matrix_questions WHERE id=? AND test_id=?",
        (question_id, test_id),
    )
    if not question_rows:
        return 0
    question = question_rows[0]
    if not bool(question["is_multiple_choice"]):
        return 0
    connection = database.connection
    changed = 0
    try:
        rows = database.rows(
            "SELECT s.attempt_id, s.score, s.response_text FROM scores s "
            "JOIN test_attempts a ON a.id=s.attempt_id "
            "WHERE a.test_id=? AND s.question_id=? AND s.response_text IS NOT NULL",
            (test_id, question_id),
        )
        affected_attempts: set[int] = set()
        for row in rows:
            response = normalize_multiple_choice_response(str(row["response_text"] or ""))
            new_score = 0.0 if response == "N" else _multiple_choice_score(question, response)
            old_score = float(row["score"]) if row["score"] is not None else None
            if (old_score is None) != (new_score is None) or (
                old_score is not None
                and new_score is not None
                and not math.isclose(old_score, new_score, rel_tol=0.0, abs_tol=1e-9)
            ):
                connection.execute(
                    "UPDATE scores SET score=? WHERE attempt_id=? AND question_id=?",
                    (new_score, row["attempt_id"], question_id),
                )
                changed += 1
            affected_attempts.add(int(row["attempt_id"]))
        for attempt_id in affected_attempts:
            total = connection.execute(
                "SELECT SUM(score) FROM scores WHERE attempt_id=?",
                (attempt_id,),
            ).fetchone()[0]
            connection.execute(
                "UPDATE test_attempts SET total_score=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (total, attempt_id),
            )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    return changed


def save_status(
    database: SubjectDatabase,
    test_id: int,
    student_id: int,
    status: str,
    *,
    commit: bool = True,
) -> None:
    if status not in RESULT_STATUSES:
        raise ResultValidationError("Deze leerlingstatus is niet toegestaan.")
    connection = database.connection
    try:
        connection.execute(
            "INSERT INTO test_attempts(test_id, student_id, status) VALUES(?, ?, ?) "
            "ON CONFLICT(test_id, student_id) DO UPDATE SET status=excluded.status, updated_at=CURRENT_TIMESTAMP",
            (test_id, student_id, status),
        )
        attempt = connection.execute(
            "SELECT id FROM test_attempts WHERE test_id=? AND student_id=?",
            (test_id, student_id),
        ).fetchone()
        if attempt and status not in {"gemaakt", "niet analyseren"}:
            connection.execute("DELETE FROM scores WHERE attempt_id=?", (attempt["id"],))
            connection.execute(
                "UPDATE test_attempts SET total_score=NULL, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (attempt["id"],),
            )
        if commit:
            connection.commit()
    except Exception:
        if commit:
            connection.rollback()
        raise
