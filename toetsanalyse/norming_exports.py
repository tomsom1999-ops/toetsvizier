from __future__ import annotations

import html
from pathlib import Path


def _number(value: float, decimals: int = 1) -> str:
    return f"{value:.{decimals}f}".replace(".", ",")


def participant_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    cleaned: list[dict[str, object]] = []
    for row in rows:
        name = str(row.get("name", "")).strip()
        if not name:
            continue
        score = float(row.get("score", 0))
        percentage = float(row.get("percentage", 0))
        grade = float(row.get("grade", 0))
        cleaned.append(
            {
                "name": name,
                "score": score,
                "percentage": percentage,
                "grade": grade,
                "status": "Voldoende" if grade >= 5.5 else "Onvoldoende",
            }
        )
    return cleaned


def build_participant_overview_html(
    test: dict[str, object],
    maximum_score: float,
    rows: list[dict[str, object]],
    method: str,
    is_finalized: bool,
) -> str:
    cleaned = participant_rows(rows)
    state = "Vastgestelde normering" if is_finalized else "Conceptnormering"
    table_rows = "".join(
        "<tr>"
        f"<td>{index}</td><td>{html.escape(str(row['name']))}</td>"
        f"<td>{_number(float(row['score']))} / {_number(maximum_score)}</td>"
        f"<td>{_number(float(row['percentage']), 0)}%</td>"
        f"<td>{_number(float(row['grade']))}</td>"
        f"<td><span class=\"badge {'good' if row['status'] == 'Voldoende' else 'bad'}\">{row['status']}</span></td>"
        "</tr>"
        for index, row in enumerate(cleaned, start=1)
    )
    test_name = html.escape(str(test.get("name", "Toets")))
    context = " | ".join(
        html.escape(str(value))
        for value in (test.get("school_year"), test.get("level"), test.get("grade_year"), test.get("period"))
        if value
    )
    return f"""<!doctype html>
<html lang="nl">
<head>
<meta charset="utf-8">
<style>
@page {{ size: A4 landscape; margin: 14mm; }}
* {{ box-sizing: border-box; }}
body {{ margin:0; background:#f7f8fc; color:#17243c; font:12px "Segoe UI", Arial, sans-serif; }}
.page {{ padding:24px; }}
h1 {{ margin:0 0 6px; font-size:25px; }}
.context {{ color:#66748e; margin-bottom:20px; }}
.cards {{ display:flex; gap:12px; margin-bottom:18px; }}
.card {{ background:#fff; border:1px solid #e4e8f1; border-radius:12px; padding:12px 16px; min-width:190px; }}
.label {{ color:#66748e; font-size:11px; margin-bottom:5px; }}
.value {{ font-size:15px; font-weight:600; }}
.table-card {{ background:#fff; border:1px solid #e4e8f1; border-radius:12px; overflow:hidden; }}
.table-title {{ font-weight:700; font-size:15px; padding:15px 16px; border-bottom:1px solid #e4e8f1; }}
table {{ width:100%; border-collapse:collapse; }}
th {{ background:#fafbfe; text-align:left; color:#66748e; font-weight:600; }}
td, th {{ padding:9px 13px; border-bottom:1px solid #edf0f6; }}
tr:last-child td {{ border-bottom:0; }}
.badge {{ padding:3px 12px; border-radius:7px; display:inline-block; }}
.good {{ background:#e9f8ee; color:#2da65a; }}
.bad {{ background:#fdf0f0; color:#e64141; }}
</style>
</head>
<body><div class="page">
<h1>Overzicht deelnemers - {test_name}</h1>
<div class="context">{context}</div>
<div class="cards">
  <div class="card"><div class="label">Normering</div><div class="value">{html.escape(state)}</div></div>
  <div class="card"><div class="label">Methode</div><div class="value">{html.escape(method)}</div></div>
  <div class="card"><div class="label">Maximumscore</div><div class="value">{_number(maximum_score)} punten</div></div>
  <div class="card"><div class="label">Deelnemers</div><div class="value">{len(cleaned)}</div></div>
</div>
<div class="table-card">
<div class="table-title">Deelnemers en cijfers</div>
<table><thead><tr><th>#</th><th>Deelnemer</th><th>Score</th><th>% van de punten gescoord</th><th>Cijfer</th><th>Status</th></tr></thead>
<tbody>{table_rows}</tbody></table>
</div></div></body></html>"""


def export_participant_overview_xlsx(
    output_path: str | Path,
    test: dict[str, object],
    maximum_score: float,
    rows: list[dict[str, object]],
    method: str,
    is_finalized: bool,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    cleaned = participant_rows(rows)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Deelnemersoverzicht"
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = f"Overzicht deelnemers - {test.get('name', 'Toets')}"
    sheet["A1"].font = Font(bold=True, size=18, color="17243C")
    sheet.merge_cells("A1:F1")
    context = " | ".join(
        str(value)
        for value in (test.get("school_year"), test.get("level"), test.get("grade_year"), test.get("period"))
        if value
    )
    sheet["A2"] = context
    sheet["A2"].font = Font(color="66748E")
    sheet.merge_cells("A2:F2")
    sheet["A4"] = "Normering"
    sheet["B4"] = "Vastgestelde normering" if is_finalized else "Conceptnormering"
    sheet["C4"] = "Methode"
    sheet["D4"] = method
    sheet["E4"] = "Maximumscore"
    sheet["F4"] = maximum_score
    for cell in sheet[4]:
        cell.fill = PatternFill("solid", fgColor="F6F8FB")
        cell.font = Font(bold=cell.column % 2 == 1, color="17243C")
    headers = ["#", "Deelnemer", "Score", "% van de punten gescoord", "Cijfer", "Status"]
    sheet.append([])
    sheet.append(headers)
    header_row = 6
    for cell in sheet[header_row]:
        cell.fill = PatternFill("solid", fgColor="17243C")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center")
    for index, row in enumerate(cleaned, start=1):
        sheet.append([index, row["name"], row["score"], row["percentage"] / 100, row["grade"], row["status"]])
    thin = Side(style="thin", color="E4E8F1")
    for row in sheet.iter_rows(min_row=header_row, max_row=header_row + len(cleaned), min_col=1, max_col=6):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center")
    for row_number in range(header_row + 1, header_row + len(cleaned) + 1):
        sheet.cell(row_number, 3).number_format = "0.0"
        sheet.cell(row_number, 4).number_format = "0%"
        sheet.cell(row_number, 5).number_format = "0.0"
        status = sheet.cell(row_number, 6)
        sufficient = status.value == "Voldoende"
        status.fill = PatternFill("solid", fgColor="E9F8EE" if sufficient else "FDF0F0")
        status.font = Font(color="2DA65A" if sufficient else "E64141")
    widths = {"A": 7, "B": 30, "C": 14, "D": 12, "E": 12, "F": 18}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A7"
    sheet.auto_filter.ref = f"A6:F{max(6, 6 + len(cleaned))}"
    workbook.save(output_path)
